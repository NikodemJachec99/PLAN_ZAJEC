import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, timezone, time as dtime, date as ddate
import math
import heapq
import re
import os
from typing import Optional, Dict, Tuple, List
from zoneinfo import ZoneInfo

# =========================
# STAŁE / KONFIG
# =========================
HOUR_HEIGHT_PX = 65         # 1h = 65px
COMPACT_RANGE = True        # przycinaj widok do zakresu zajęć (+/- 15 min)
TZ_WA = ZoneInfo("Europe/Warsaw")

MAIN_PLAN_FILE = "plan_zajec.xlsx"
PRACTICAL_CANDIDATES = [
    "Pi_s_II_sem.-zimowy_26.09.2025 (1).xlsx",
    "/mnt/data/Pi_s_II_sem.-zimowy_26.09.2025 (1).xlsx",
    "praktyki.xlsx",
]

# =========================
# USTAWIENIA STRONY
# =========================
st.set_page_config(page_title="Plan Zajęć ❤️", page_icon="📅", layout="centered")

# --- AUTO-ODŚWIEŻANIE (60 s) ---
try:
    from streamlit_autorefresh import st_autorefresh
    st_autorefresh(interval=60 * 1000, key="autorefresh-60s")
except Exception:
    st.markdown("<script>setTimeout(()=>window.location.reload(), 60000);</script>", unsafe_allow_html=True)

# =========================
# POMOCNICZE
# =========================
def _first_existing(paths):
    for p in paths:
        if os.path.exists(p):
            return p
    return None

def to_minutes(t: dtime) -> int:
    return t.hour * 60 + t.minute

def _hex_to_rgb_tuple(hx: Optional[str]):
    if not hx or not isinstance(hx, str) or not hx.startswith("#") or len(hx) != 7:
        return None
    try:
        r = int(hx[1:3], 16); g = int(hx[3:5], 16); b = int(hx[5:7], 16)
        return (r,g,b)
    except Exception:
        return None

# =========================
# PARSER GŁÓWNEGO PLANU
# =========================
@st.cache_data(ttl=600)
def load_data(file_path: str) -> pd.DataFrame:
    df = pd.read_excel(file_path, header=3)

    df.columns = [
        'date', 'day_of_week', 'start_time', 'end_time', 'subject', 'type',
        'degree', 'first_name', 'last_name', 'room', 'field_year', 'group',
        'info_combined', 'additional_info'
    ] + [f'unnamed_{i}' for i in range(len(df.columns) - 14)]

    df = df[['date', 'day_of_week', 'start_time', 'end_time', 'subject',
             'type', 'degree', 'first_name', 'last_name', 'room', 'group']].copy()
    df.dropna(subset=['date'], inplace=True)
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df.dropna(subset=['date'], inplace=True)

    df['instructor'] = (
        df['degree'].fillna('') + ' ' + df['first_name'].fillna('') + ' ' + df['last_name'].fillna('')
    ).str.strip()
    df['group'] = df['group'].fillna('---').astype(str)

    df['start_time_obj'] = pd.to_datetime(df['start_time'], format='%H:%M:%S', errors='coerce').dt.time
    df['end_time_obj']   = pd.to_datetime(df['end_time'],   format='%H:%M:%S', errors='coerce').dt.time
    df['start_time'] = df['start_time_obj'].apply(lambda x: x.strftime('%H:%M') if pd.notnull(x) else 'Błąd')
    df['end_time']   = df['end_time_obj'].apply(  lambda x: x.strftime('%H:%M') if pd.notnull(x) else 'Błąd')

    df.sort_values(by=['date', 'start_time_obj'], inplace=True)
    return df

# =========================
# PRAKTYKI — PARSER „STARY” (pandas), na wypadek braku arkusza 'grafik'
# =========================
def _month_from_label_old(lbl: str):
    PL = {'STYCZEŃ':1,'LUTY':2,'MARZEC':3,'KWIECIEŃ':4,'MAJ':5,'CZERWIEC':6,
          'LIPIEC':7,'SIERPIEŃ':8,'WRZESIEŃ':9,'PAŹDZIERNIK':10,'LISTOPAD':11,'GRUDZIEŃ':12}
    if not isinstance(lbl, str): return None
    return PL.get(lbl.strip().upper())

def _build_col_date_table_old(df: pd.DataFrame):
    """Mapuje kolumny -> daty na podstawie: wiersz 3 (miesiąc), wiersz 5 (dzień) i nagłówka z latami."""
    header_text = str(df.iat[1,0]) if df.shape[0] > 1 and pd.notna(df.iat[1,0]) else ""
    m = re.search(r'(20\d{2})/(20\d{2})', header_text)
    start_year, end_year = (int(m.group(1)), int(m.group(2))) if m else (datetime.now().year, datetime.now().year+1)

    current_month = None
    col_date = {}
    for c in range(1, df.shape[1]):
        mcell = df.iat[2, c] if df.shape[0] > 2 else None
        mm = _month_from_label_old(mcell)
        if mm: current_month = mm
        d = df.iat[4, c] if df.shape[0] > 4 else None
        if pd.notna(d) and current_month is not None:
            try:
                day = int(float(d))
                year = start_year if current_month >= 10 else end_year
                col_date[c] = pd.to_datetime(f"{year}-{current_month:02d}-{day:02d}").date()
            except Exception:
                pass
    return col_date

def _find_group_row_old(df: pd.DataFrame, label='11'):
    """Znajdź wiersz, którego pierwsza kolumna to '11' (obsługuje 11/11.0/'11')."""
    want = str(label)
    for i in range(df.shape[0]):
        v = df.iat[i, 0]
        if pd.isna(v):
            continue
        if isinstance(v, (int, float)) and float(v).is_integer():
            s = str(int(v))
        else:
            s = str(v).strip()
            if s.endswith('.0') and s[:-2].isdigit():
                s = s[:-2]
        if s == want:
            return i
    return None

def _build_code_info_old(df: pd.DataFrame):
    """Buduje mapę: kod -> {subject, ward, location, teacher, default_time} z legendy (wiersze 22..38)."""
    code_info = {}
    time_re = re.compile(r'(\d{1,2}:\d{2}).*?(\d{1,2}:\d{2})', re.I)

    def looks_location(s: str) -> bool:
        s2 = s.lower()
        return (' ul.' in s2) or ('szpital' in s2) or ('przychodnia' in s2) or (',' in s2 and len(s) > 10)

    def looks_teacher(s: str) -> bool:
        return (len(s) > 8 and (' ' in s or '-' in s)) and not s.isupper()

    for r in range(21, min(40, df.shape[0])):  # 22..38 (0-index)
        cells = []
        for c in range(df.shape[1]):
            v = df.iat[r, c]
            if isinstance(v, str) and v.strip():
                cells.append(v.strip())
            elif isinstance(v, (int, float)) and not pd.isna(v):
                cells.append(str(int(v)) if float(v).is_integer() else str(v))
        if not cells:
            continue

        raw_code = None
        if df.shape[1] > 1:
            v = df.iat[r, 1]
            if isinstance(v, (int, float)) and not pd.isna(v):
                raw_code = str(int(v)) if float(v).is_integer() else str(v)
            elif isinstance(v, str):
                raw_code = v.strip()
        if not raw_code:
            m = re.search(r'\b([A-Za-z]{1,4}\d{0,3}|\d{1,3})\b', ' '.join(cells))
            raw_code = m.group(1) if m else None
        if not raw_code:
            continue

        code = raw_code.upper()
        subject = ''
        ward = ''
        location = ''
        teacher = ''
        default_time = None

        for s in cells:
            m = time_re.search(s)
            if m:
                default_time = (m.group(1), m.group(2))
                break

        long_texts = [s for s in cells if len(s) > 12 and not time_re.search(s)]
        if long_texts:
            subject = max(long_texts, key=len)

        wards = [s for s in cells if re.fullmatch(r'[A-Z]{2,4}', s)]
        if wards:
            for pref in ('POZ','CSM','SOR','ZOL','USK','UO'):
                if pref in wards:
                    ward = pref; break
            if not ward:
                ward = wards[0]

        locs = [s for s in cells if looks_location(s)]
        if locs:
            location = max(locs, key=len)

        teachers = [s for s in cells if looks_teacher(s)]
        if teachers:
            teacher = teachers[-1]

        code_info[code] = {
            'subject': subject or code,
            'ward': ward,
            'location': location,
            'teacher': teacher,
            'default_time': default_time
        }

    code_info.setdefault('CSM', {'subject':'Centrum Symulacji Medycznych','ward':'CSM',
                                 'location':'','teacher':'','default_time':None})
    return code_info

@st.cache_data(ttl=600)
def load_practicals_group11_old(file_path: str) -> pd.DataFrame:
    raw = pd.read_excel(file_path, sheet_name="grafik", header=None)

    col_date = _build_col_date_table_old(raw)
    grp_row = _find_group_row_old(raw, '11')
    if grp_row is None:
        grp_row = 10 if raw.shape[0] > 10 else None
    if grp_row is None:
        return pd.DataFrame(columns=["date","subject","instructor","room","group",
                                     "start_time_obj","end_time_obj","start_time","end_time","cell_color"])

    code_info = _build_code_info_old(raw)

    time_range_re = re.compile(r'(\d{1,2}[:.]\d{2})\s*[-–]\s*(\d{1,2}[:.]\d{2})')

    def _parse_hhmm_forgiving(s: str):
        s = str(s).strip()
        m = re.match(r'^(\d{1,2})[:.](\d{2})$', s)
        if not m:
            return None
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return dtime(h, mi)
        return None

    rows_out = []
    for c, the_date in col_date.items():
        cell = raw.iat[grp_row, c]
        if pd.isna(cell):
            continue
        pieces = re.split(r'[;\n]+', str(cell))
        for piece in (p.strip() for p in pieces if p and p.strip()):
            times = time_range_re.findall(piece)
            core = time_range_re.sub('', piece)
            core = re.sub(r'(?i)godz\.\s*', '', core).strip(" -–•.,;")

            m_code = re.search(r'\b([A-Za-z]{1,4}\d{0,3}|\d{1,3})\b', core)
            code = m_code.group(1).upper() if m_code else None

            meta = code_info.get(code, {})
            subject = meta.get('subject') or (core if core else "Praktyki")
            ward = meta.get('ward','')
            location = meta.get('location','')
            teacher = meta.get('teacher','')
            room = " • ".join([x for x in [ward, location] if x]) or code or ""

            if times:
                for s, e in times:
                    stime = _parse_hhmm_forgiving(s)
                    etime = _parse_hhmm_forgiving(e)
                    if not stime or not etime:
                        continue
                    rows_out.append({
                        "date": pd.to_datetime(the_date),
                        "subject": subject,
                        "instructor": teacher or "",
                        "room": room or "",
                        "group": "11",
                        "start_time_obj": stime,
                        "end_time_obj": etime,
                        "start_time": stime.strftime("%H:%M"),
                        "end_time": etime.strftime("%H:%M"),
                        "cell_color": None
                    })
            else:
                s, e = (meta.get('default_time') or ('07:30','15:00'))
                stime = _parse_hhmm_forgiving(s)
                etime = _parse_hhmm_forgiving(e)
                if stime and etime:
                    rows_out.append({
                        "date": pd.to_datetime(the_date),
                        "subject": subject,
                        "instructor": teacher or "",
                        "room": room or "",
                        "group": "11",
                        "start_time_obj": stime,
                        "end_time_obj": etime,
                        "start_time": stime.strftime("%H:%M"),
                        "end_time": etime.strftime("%H:%M"),
                        "cell_color": None
                    })

    out = pd.DataFrame(rows_out)
    if not out.empty:
        out.sort_values(by=["date", "start_time_obj"], inplace=True)
        out.reset_index(drop=True, inplace=True)
    return out

# =========================
# PRAKTYKI — PARSER „NOWY” (openpyxl + kolory) dla pliku z arkuszem 'grafik'
# =========================
MONTHS_PL = {
    'STYCZEŃ':1,'STYCZEN':1,'LUTY':2,'MARZEC':3,'KWIECIEŃ':4,'KWIECIEN':4,
    'MAJ':5,'CZERWIEC':6,'LIPIEC':7,'SIERPIEŃ':8,'SIERPIEN':8,
    'WRZESIEŃ':9,'WRZESIEN':9,'PAŹDZIERNIK':10,'PAZDZIERNIK':10,
    'LISTOPAD':11,'GRUDZIEŃ':12,'GRUDZIEN':12
}

def _normalize_str(x) -> str:
    if x is None: return ""
    if isinstance(x, float) and x.is_integer():
        return str(int(x))
    return str(x).strip()

def _find_years_from_header(ws) -> Tuple[int, int]:
    for r in range(1, min(6, ws.max_row)+1):
        row_text = " ".join(_normalize_str(ws.cell(r,c).value) for c in range(1, min(8, ws.max_column)+1))
        m = re.search(r'(20\d{2})\s*[/\-]\s*(20\d{2})', row_text)
        if m:
            return int(m.group(1)), int(m.group(2))
    today = datetime.now(TZ_WA).date()
    return (today.year, today.year+1) if today.month >= 9 else (today.year-1, today.year)

def _find_header_rows(ws) -> Tuple[Optional[int], Optional[int]]:
    month_row = None
    day_row = None
    for r in range(1, min(12, ws.max_row)+1):
        texts = [_normalize_str(ws.cell(r, c).value).upper() for c in range(1, min(40, ws.max_column)+1)]
        if any(t in MONTHS_PL for t in texts if t):
            month_row = r
            break
    for r in range(1, min(15, ws.max_row)+1):
        nums = 0
        for c in range(1, min(60, ws.max_column)+1):
            v = ws.cell(r, c).value
            if v is None: continue
            if isinstance(v, float) and v.is_integer(): v = int(v)
            if isinstance(v, int) and 1 <= v <= 31:
                nums += 1
        if nums >= 5:
            day_row = r
            break
    return month_row, day_row

def _month_from_label(lbl: str) -> Optional[int]:
    if not isinstance(lbl, str): return None
    s = lbl.strip().upper()
    return MONTHS_PL.get(s)

def _build_col_date_table_xl(ws) -> Dict[int, ddate]:
    start_year, end_year = _find_years_from_header(ws)
    month_row, day_row = _find_header_rows(ws)
    if month_row is None: month_row = 3
    if day_row is None: day_row = 5

    col_date: Dict[int, ddate] = {}
    current_month = None
    for c in range(2, ws.max_column+1):
        mcell = ws.cell(month_row, c).value
        mm = _month_from_label(_normalize_str(mcell))
        if mm:
            current_month = mm
        dcell = ws.cell(day_row, c).value
        if dcell is None or current_month is None:
            continue
        try:
            if isinstance(dcell, float) and dcell.is_integer():
                day = int(dcell)
            elif isinstance(dcell, int):
                day = dcell
            else:
                day = int(str(dcell).strip().split()[0])
        except Exception:
            continue
        if not (1 <= day <= 31):
            continue
        year = start_year if current_month >= 9 else end_year
        try:
            dt = ddate(year, current_month, day)
        except Exception:
            continue
        col_date[c] = dt
    return col_date

def _find_group_row_xl(ws, group_label='11') -> Optional[int]:
    want = str(group_label)
    for r in range(1, min(200, ws.max_row)+1):
        v = ws.cell(r,1).value
        if v is None:
            continue
        s = _normalize_str(v)
        if s.endswith('.0') and s[:-2].isdigit():
            s = s[:-2]
        if s == want:
            return r
    return None

def _parse_forgiving_time(s: str) -> Optional[dtime]:
    s = _normalize_str(s)
    m = re.match(r'^\s*(\d{1,2})[:.](\d{2})\s*$', s)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if 0 <= h <= 23 and 0 <= mi <= 59:
        return dtime(h, mi)
    return None

def _extract_color(cell) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    fg = getattr(fill, "fgColor", None) if fill else None
    rgb = getattr(fg, "rgb", None) if fg else None
    if not rgb:
        return None
    if len(rgb) == 8:
        return "#" + rgb[-6:]
    if len(rgb) == 6:
        return "#" + rgb
    return None

def _build_code_info_xl(ws) -> Dict[str, dict]:
    code_info = {}
    time_re = re.compile(r'(\d{1,2}[:.]\d{2}).*?(\d{1,2}[:.]\d{2})')

    def looks_location(s: str) -> bool:
        s2 = s.lower()
        return (' ul.' in s2) or ('szpital' in s2) or ('przychodnia' in s2) or (',' in s2 and len(s) > 10)

    def looks_teacher(s: str) -> bool:
        return (len(s) > 8 and (' ' in s or '-' in s)) and not s.isupper()

    for r in range(18, min(70, ws.max_row)+1):
        cells = [_normalize_str(ws.cell(r,c).value) for c in range(1, ws.max_column+1)]
        cells = [s for s in cells if s]
        if not cells:
            continue

        raw_code = _normalize_str(ws.cell(r,2).value)
        if not raw_code:
            m = re.search(r'\b([A-Za-z]{1,4}\d{0,3}|\d{1,3})\b', " ".join(cells))
            raw_code = m.group(1) if m else None
        if not raw_code:
            continue

        code = raw_code.upper()
        subject, ward, location, teacher, default_time = '', '', '', '', None

        for s in cells:
            m = time_re.search(s)
            if m:
                default_time = (m.group(1).replace('.',':'), m.group(2).replace('.',':'))
                break

        long_texts = [s for s in cells if len(s) > 12 and not time_re.search(s)]
        if long_texts:
            subject = max(long_texts, key=len)

        wards = [s for s in cells if re.fullmatch(r'[A-Z]{2,4}', s)]
        if wards:
            for pref in ('POZ','CSM','SOR','ZOL','USK','UO'):
                if pref in wards:
                    ward = pref; break
            if not ward:
                ward = wards[0]

        locs = [s for s in cells if looks_location(s)]
        if locs:
            location = max(locs, key=len)

        teachers = [s for s in cells if looks_teacher(s)]
        if teachers:
            teacher = teachers[-1]

        code_info[code] = {
            'subject': subject or code,
            'ward': ward,
            'location': location,
            'teacher': teacher,
            'default_time': default_time
        }

    code_info.setdefault('CSM', {'subject':'Centrum Symulacji Medycznych','ward':'CSM',
                                 'location':'','teacher':'','default_time':None})
    return code_info

@st.cache_data(ttl=600)
def load_practicals_group11_uploaded(file_path: str) -> pd.DataFrame:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    if "grafik" not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb["grafik"]

    col_date = _build_col_date_table_xl(ws)
    grp_row = _find_group_row_xl(ws, '11') or 11
    code_info = _build_code_info_xl(ws)

    time_range_re = re.compile(r'(\d{1,2}[:.]\d{2})\s*[-–]\s*(\d{1,2}[:.]\d{2})')
    rows_out: List[dict] = []

    for c, the_date in col_date.items():
        cell = ws.cell(grp_row, c)
        raw = _normalize_str(cell.value)
        if not raw:
            continue
        cell_color = _extract_color(cell)

        pieces = re.split(r'[;\n]+', raw)
        for piece in (p.strip() for p in pieces if p and p.strip()):
            times = time_range_re.findall(piece.replace('godz.', '').replace('godz', ''))
            core = time_range_re.sub('', piece)
            core = re.sub(r'(?i)godz\.\s*', '', core).strip(" -–•.,;")

            m_code = re.search(r'\b([A-Za-z]{1,4}\d{0,3}|\d{1,3})\b', core)
            code = m_code.group(1).upper() if m_code else None

            meta = code_info.get(code or "", {})
            subject = meta.get('subject') or (core if core else "Praktyki")
            ward = meta.get('ward','')
            location = meta.get('location','')
            teacher = meta.get('teacher','')
            room = " • ".join([x for x in [ward, location] if x]) or (code or "")

            if times:
                for s,e in times:
                    stime = _parse_forgiving_time(s.replace('.',':'))
                    etime = _parse_forgiving_time(e.replace('.',':'))
                    if not stime or not etime:
                        continue
                    rows_out.append({
                        "date": pd.to_datetime(the_date),
                        "subject": subject,
                        "instructor": teacher or "",
                        "room": room or "",
                        "group": "11",
                        "start_time_obj": stime,
                        "end_time_obj": etime,
                        "start_time": stime.strftime("%H:%M"),
                        "end_time": etime.strftime("%H:%M"),
                        "cell_color": cell_color
                    })
            else:
                s, e = (meta.get('default_time') or ('07:30','15:00'))
                stime = _parse_forgiving_time(s.replace('.',':'))
                etime = _parse_forgiving_time(e.replace('.',':'))
                if stime and etime:
                    rows_out.append({
                        "date": pd.to_datetime(the_date),
                        "subject": subject,
                        "instructor": teacher or "",
                        "room": room or "",
                        "group": "11",
                        "start_time_obj": stime,
                        "end_time_obj": etime,
                        "start_time": stime.strftime("%H:%M"),
                        "end_time": etime.strftime("%H:%M"),
                        "cell_color": cell_color
                    })

    out = pd.DataFrame(rows_out)
    if not out.empty:
        out.sort_values(by=["date", "start_time_obj"], inplace=True)
        out.reset_index(drop=True, inplace=True)
    return out

# =========================
# WYBÓR PARSERA PRAKTYK
# =========================
@st.cache_data(ttl=600)
def load_practicals_auto(file_path: str) -> pd.DataFrame:
    """Wybiera parser: jeśli w xlsx jest arkusz 'grafik' -> nowy (openpyxl + kolory).
       Jeśli nie da rady lub wyjdzie pusto, próbuje parsera starego."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        use_new = "grafik" in wb.sheetnames
    except Exception:
        use_new = False

    df = pd.DataFrame()
    if use_new:
        try:
            df = load_practicals_group11_uploaded(file_path)
        except Exception:
            df = pd.DataFrame()

    if df.empty:
        # fallback do „starego” parsera (pandas), też szuka 'grafik'
        try:
            df = load_practicals_group11_old(file_path)
        except Exception:
            df = pd.DataFrame()

    # Upewnij się, że są kolumny zgodne z resztą UI
    if not df.empty:
        for col in ["cell_color"]:
            if col not in df.columns:
                df[col] = None
    return df

# =========================
# STYLE
# =========================
st.markdown("""
<style>
  html { font-size: 14px; }
  body, .stApp { line-height: 1.25; }
  .stApp > header { background-color: transparent; }
  .main .block-container { padding: 0.5rem 0.6rem 3rem 0.6rem; }
  h1 { text-align:center; color:#1a202c; margin-bottom:0.6rem; font-size:1.35rem; }
  .week-range { text-align:center; font-size:1.05rem; font-weight:600; color:#2d3748; margin:0.2rem 0 0.6rem; }

  .stButton>button {
    padding: 0.25rem 0.5rem !important;
    font-size: 0.85rem !important;
    line-height: 1.1 !important;
    border-radius: 8px !important;
  }

  .day-layout { display:grid; grid-template-columns:70px 1fr; gap:0.6rem; align-items:start; }

  .timeline-rail { position:sticky; top:0; width:70px; border-right:2px solid #e2e8f0; }
  .timeline-rail-inner { position:relative; height:var(--day-height, 720px); }
  .tick { position:absolute; left:0; right:0; border-top:1px dashed #e2e8f0; }
  .tick-label { position:absolute; left:0; width:60px; text-align:right; font-size:0.72rem; color:#94a3b8; transform:translateY(-50%); padding-right:4px; }

  .calendar-canvas { position:relative; min-height:var(--day-height, 720px); border-left:2px solid #e2e8f0; }
  .event {
    position:absolute; box-sizing:border-box; padding:6px 8px;
    background:#0ea5e910; border:1px solid #38bdf8; border-radius:10px;
    overflow:hidden; box-shadow:0 1px 2px rgba(0,0,0,.05);
  }
  .event .title { font-weight:700; color:#0f172a; margin-bottom:1px; font-size:0.92rem; }
  .event .meta { font-size:.72rem; color:#334155; line-height:1.15; }

  .now-line-wide { position:absolute; left:0; right:0; border-top:2px solid #ef4444; z-index:3; }
  .now-badge { position:absolute; right:6px; transform:translateY(-100%); font-size:.7rem; color:#ef4444; z-index:4; background:transparent; }

  @media (max-width: 640px) {
    html { font-size: 13px; }
    .day-layout { grid-template-columns:56px 1fr; gap:0.45rem; }
    .timeline-rail { width:56px; }
    .tick-label { width:48px; font-size:.68rem; }
    .event { padding:5px 7px; border-radius:8px; }
    .event .title { font-size:.88rem; }
    .event .meta { font-size:.68rem; }
    .stButton>button { padding:0.22rem 0.45rem !important; font-size:.82rem !important; }
  }
</style>
""", unsafe_allow_html=True)

# =========================
# KALENDARZ / UI
# =========================
def assign_columns_and_clusters(evts):
    result = []; active = []; free_cols = []; next_col = 0
    clusters, current_cluster = [], []
    for idx, ev in enumerate(evts):
        while active and active[0][0] <= ev["start_min"]:
            _, col_finished, _ = heapq.heappop(active)
            free_cols.append(col_finished); free_cols.sort()
        if not active and current_cluster:
            clusters.append(current_cluster); current_cluster = []
        col = free_cols.pop(0) if free_cols else next_col
        if col == next_col: next_col += 1
        heapq.heappush(active, (ev["end_min"], col, idx))
        result.append({**ev, "col": col, "cluster_id": -1})
        current_cluster.append((idx, ev["start_min"], ev["end_min"], col))
    if current_cluster: clusters.append(current_cluster)
    cluster_cols = {}
    for c_id, items in enumerate(clusters):
        points = []
        for _, s, e, _ in items:
            points.append((s, 1)); points.append((e, -1))
        points.sort(key=lambda x: (x[0], -x[1]))
        cur = peak = 0
        for _, d in points:
            cur += d; peak = max(peak, cur)
        cluster_cols[c_id] = max(1, peak)
        for idx, *_ in items:
            result[idx]["cluster_id"] = c_id
    return result, cluster_cols

try:
    # Główny plan
    df_main = load_data(MAIN_PLAN_FILE)

    # Praktyki (grupa 11): autodetekcja parsera
    df_pr = pd.DataFrame()
    prat_file = _first_existing(PRACTICAL_CANDIDATES)
    if prat_file:
        try:
            df_pr = load_practicals_auto(prat_file)
        except Exception:
            df_pr = pd.DataFrame()

    st.title("Plan Zajęć ❤️")

    # Czas Warszawy
    now_dt = datetime.now(timezone.utc).astimezone(TZ_WA)
    today = now_dt.date()

    # Stan sesji
    if 'current_week_start' not in st.session_state:
        st.session_state.current_week_start = today - timedelta(days=today.weekday())
    if 'selected_day_index' not in st.session_state:
        st.session_state.selected_day_index = today.weekday()

    # Nawigacja tygodnia
    week_start = st.session_state.current_week_start
    week_end = week_start + timedelta(days=6)

    nav_cols = st.columns([1, 1, 4, 1])
    if nav_cols[0].button("⬅️ Poprzedni", use_container_width=True):
        st.session_state.current_week_start -= timedelta(days=7)
        st.rerun()
    if nav_cols[1].button("📍 Dziś", use_container_width=True):
        st.session_state.current_week_start = today - timedelta(days=today.weekday())
        st.session_state.selected_day_index = today.weekday()
        st.rerun()
    nav_cols[2].markdown(f"<div class='week-range'>{week_start.strftime('%d.%m')} – {week_end.strftime('%d.%m.%Y')}</div>", unsafe_allow_html=True)
    if nav_cols[3].button("Następny ➡️", use_container_width=True):
        st.session_state.current_week_start += timedelta(days=7)
        st.rerun()

    # Checkbox „Grupy Magdalenki” – działa WYŁĄCZNIE na df_main
    filter_magdalenki = st.checkbox("**:red[Grupy Magdalenki]**", value=False)

    df_main_src = df_main
    if filter_magdalenki:
        def _is_whole_year(g: str) -> bool:
            s = (g or "").strip().lower()
            return (s in {"---", "rok", "wszyscy", "all", "year"}) or ("rok" in s) or ("wsz" in s)

        def _keep_row(grp: str) -> bool:
            s = (grp or "").strip().lower()
            if _is_whole_year(s): return True
            if any(ch.isdigit() for ch in s):      # numeryczne → tylko '11'
                return s == "11" or s.startswith("11")
            return s == "d" or s.startswith("d")   # literowe → tylko 'd'

        df_main_src = df_main[df_main["group"].apply(_keep_row)]

    # Doklej praktyki (grupa 11) niezależnie od filtra
    if not df_pr.empty:
        df_all = pd.concat([df_main_src, df_pr], ignore_index=True, sort=False)
    else:
        df_all = df_main_src
    if not df_all.empty:
        df_all.sort_values(by=['date', 'start_time_obj'], inplace=True)

    # Skala pionowa
    PX_PER_MIN = HOUR_HEIGHT_PX / 60.0

    # Zakładki dni (Pon–Pt)
    days_of_week_pl = ["Pon", "Wt", "Śr", "Czw", "Pt", "Sob", "Niedz"]
    visible_offsets = [0, 1, 2, 3, 4]
    day_tabs = st.columns(len(visible_offsets))
    for i, off in enumerate(visible_offsets):
        current_day_date = week_start + timedelta(days=off)
        if day_tabs[i].button(f"{days_of_week_pl[off]} {current_day_date.day}", use_container_width=True):
            st.session_state.selected_day_index = off
            st.rerun()

    # Dane dnia
    selected_day_date = week_start + timedelta(days=st.session_state.selected_day_index)
    day_events = df_all[df_all['date'].dt.date == selected_day_date]
    st.markdown(f"### {selected_day_date.strftime('%A, %d.%m.%Y')}")

    # OŚ CZASU
    base_start, base_end = dtime(7, 0), dtime(21, 0)
    base_start_m, base_end_m = to_minutes(base_start), to_minutes(base_end)

    if not day_events.empty and COMPACT_RANGE:
        ev_start_m = min(day_events['start_time_obj'].dropna().map(to_minutes))
        ev_end_m   = max(day_events['end_time_obj'].dropna().map(to_minutes))
        start_m = max(base_start_m, int(math.floor((ev_start_m - 15) / 60) * 60))
        end_m   = min(base_end_m,  int(math.ceil((ev_end_m + 15) / 60) * 60))
    else:
        if not day_events.empty:
            ev_min = day_events['start_time_obj'].dropna()
            ev_max = day_events['end_time_obj'].dropna()
            start_m = min([base_start_m] + ([min(map(to_minutes, ev_min))] if len(ev_min) else []))
            end_m   = max([base_end_m]   + ([max(map(to_minutes, ev_max))] if len(ev_max) else []))
        else:
            start_m, end_m = base_start_m, base_end_m

    duration = max(60, end_m - start_m)
    height_px = duration * PX_PER_MIN

    # Godzinowe ticki
    first_tick_h = math.ceil(start_m / 60)
    last_tick_h  = math.floor(end_m / 60)
    ticks_html = []
    for h in range(first_tick_h, last_tick_h + 1):
        top = (h * 60 - start_m) * PX_PER_MIN
        ticks_html.append(f"<div class='tick' style='top:{top:.2f}px'></div>")
        ticks_html.append(f"<div class='tick-label' style='top:{top:.2f}px'>{h:02d}:00</div>")

    # Eventy
    events = []
    for _, e in day_events.iterrows():
        if pd.isna(e['start_time_obj']) or pd.isna(e['end_time_obj']):
            continue
        events.append({
            "start_min": to_minutes(e['start_time_obj']),
            "end_min": to_minutes(e['end_time_obj']),
            "subject": e.get("subject",""),
            "instructor": e.get("instructor",""),
            "room": e.get("room",""),
            "group": e.get("group",""),
            "start_str": e.get("start_time",""),
            "end_str": e.get("end_time",""),
            "cell_color": e.get("cell_color", None),
        })
    events.sort(key=lambda x: (x["start_min"], x["end_min"]))

    # Kolumny + klastry
    positioned, cluster_cols = assign_columns_and_clusters(events)

    # Render
    PX = PX_PER_MIN
    events_html_parts = []
    for ev in positioned:
        total_cols = max(1, cluster_cols.get(ev["cluster_id"], 1))
        width_pct = 100 / total_cols
        left_pct = ev["col"] * width_pct
        top = (ev["start_min"] - start_m) * PX
        height = max(34, (ev["end_min"] - ev["start_min"]) * PX)

        style_extra = ""
        if ev.get("cell_color"):
            rgb = _hex_to_rgb_tuple(ev["cell_color"])
            if rgb:
                r,g,b = rgb
                style_extra = f"background: rgba({r},{g},{b},.12); border-color: rgb({r},{g},{b});"

        part = (
            f"<div class='event' style='top:{top:.2f}px;height:{height:.2f}px;"
            f"left:calc({left_pct}% + 2px);width:calc({width_pct}% - 6px);{style_extra}'>"
            f"<div class='title'>{ev['subject']}</div>"
            f"<div class='meta'>{ev['start_str']}–{ev['end_str']} • Sala/Oddz: {ev['room']} • Gr {ev['group']}<br>{ev['instructor']}</div>"
            f"</div>"
        )
        events_html_parts.append(part)
    events_html = "".join(events_html_parts)

    # Linia TERAZ
    now_wide_html = ""
    if selected_day_date == today:
        now_dt_line = datetime.now(timezone.utc).astimezone(TZ_WA)
        now_m = now_dt_line.hour * 60 + now_dt_line.minute
        top_now = max(0, min(height_px, (now_m - start_m) * PX_PER_MIN))
        now_wide_html = (
            f"<div class='now-line-wide' style='top:{top_now:.2f}px'></div>"
            f"<div class='now-badge' style='top:{top_now:.2f}px'>Teraz {now_dt_line.strftime('%H:%M')}</div>"
        )

    # Layout
    day_layout_html = (
        f"<div class='day-layout' style='--day-height:{height_px:.2f}px'>"
        f"<div class='timeline-rail'><div class='timeline-rail-inner' style='height:{height_px:.2f}px'>{''.join(ticks_html)}</div></div>"
        f"<div class='calendar-canvas' style='min-height:{height_px:.2f}px'>{now_wide_html}{events_html if events_html else '<div style=\"padding:8px;color:#64748b;\">Brak zajęć</div>'}</div>"
        f"</div>"
    )
    st.markdown(day_layout_html, unsafe_allow_html=True)

except FileNotFoundError:
    st.error("Nie znaleziono pliku `plan_zajec.xlsx` lub arkusza praktyk.")
except Exception as e:
    st.error(f"Wystąpił nieoczekiwany błąd: {e}")
